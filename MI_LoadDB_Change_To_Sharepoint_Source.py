
import datetime
from dateutil.relativedelta import relativedelta 
import calendar 
import pandas as pd
import openpyxl
from openpyxl import Workbook,load_workbook
from pathlib import Path
import pyodbc
# from azure.storage.blob import BlobServiceClient, BlobClient
import io
from office365.sharepoint.client_context import ClientContext
from office365.runtime.auth.user_credential import UserCredential
import io
import openpyxl
from io import BytesIO
import configparser


config = configparser.ConfigParser()
config.read('configloc.ini')

configpath = config.get('CONFIG','path')
config.read(configpath)

sharepoint_site_url = config.get('Sharepoint','sharepoint_site_url')
username = config.get('Sharepoint','username')
password = config.get('Sharepoint','password')

server = config.get('archDB','toserver')
port = config.get('archDB','toport')
database = config.get('archDB','todatabase')
dbusername = config.get('archDB','tousername')
dbpassword = config.get('archDB','topassword')

driver= '{ODBC Driver 17 for SQL Server}'

def YearMonth (month):
    if month <10 :
        return '0'+str(month)
    else:
        return str(month)
    

last_month=datetime.datetime.today()-relativedelta(months=1)
yyyymm = last_month.strftime('%Y%m')
yyyy_mm = last_month.strftime('%Y_%m')

print("hello" + yyyymm)

# sharepoint_site_url = "https://convoy.sharepoint.com/sites/BI/"
sharepoint_folder_url  = f"Shared%20Documents/Shared/01_MI Reports/Raw/Service Management Complete/{yyyymm}"
# username = "szekiat.pua@convoy.com.hk"
# password = "Convoy0112$"

ctx = ClientContext(sharepoint_site_url).with_credentials(UserCredential(username, password))

def download_excel_file(folder_url, file_pattern):
    folder = ctx.web.get_folder_by_server_relative_url(folder_url)
    files = folder.files
    ctx.load(files)
    ctx.execute_query()

    matching_file = None
    for file in files:
        if file.name.startswith(file_pattern):
            matching_file = file
            break

    if matching_file:
        print(f"Found file: {matching_file.name}")
        file_content = BytesIO()
        matching_file.download(file_content).execute_query()
        file_content.seek(0)
        return openpyxl.load_workbook(file_content, data_only=True)
    else:
        print(f"No file found matching pattern '{file_pattern}' in folder '{folder_url}'")
        return None
    
mgmt_folder_url = f"Shared%20Documents/Shared/01_MI Reports/Raw/Service Overview/{yyyymm}"
operations_folder_url = f"Shared%20Documents/Shared/01_MI Reports/Raw/OWM Operation Overview/{yyyymm}"

mgmt_wb_obj = download_excel_file(mgmt_folder_url, "MI Serv Mgmt Overview")
operation_wb_obj = download_excel_file(operations_folder_url, "MI Operations Overview")

if mgmt_wb_obj and 'CFS' in mgmt_wb_obj.sheetnames:
    CFS_sheet = mgmt_wb_obj['CFS']
    OP_sheet = mgmt_wb_obj["OP"]
    
if operation_wb_obj and 'OWM' in operation_wb_obj.sheetnames:
    OWM_operation_sheet = operation_wb_obj['OWM']

#Edit this 
blob_path = "report/"
serv_blob_path = blob_path +"SCR0226/"
persis_blob_path = blob_path +"SCR0336/"

folder_path_raw = serv_blob_path + 'raw/'+yyyymm+'/'
persis_folder_path_raw = persis_blob_path + yyyymm+'/'

year =str(last_month.year)
tar_month = last_month.month
 
#################################################################################
saleforce_start_column = 5
operation_start_column = 5 
operation_end_column = operation_start_column + tar_month -1

saleforce_end_column = saleforce_start_column + tar_month -1
top3_start_row = 6
top3_end_row = 8
sales_force_table_name = 'Dwh_Mi_OperationOverview_SalesforceManual'                           #'MI_OPERATION_OVERVIEW_SALESFORCE_MANUAL'
top3_table_name = 'Dwh_Mi_OperationOverview_ReportTop3'                                  #'MI_OPERATION_OVERVIEW_REPORT_Top3'
sod_table_name = 'Dwh_Mi_OperationOverview_SodManual'                                   #'MI_OPERATION_OVERVIEW_SOD_MANUAL'

#CFS
channel = "'CFS'"
#saleforce
month_pos = 12
ce_email_case_pos = 27
ce_counter_case_pos =28
ce_hotline_case_pos =29
ce_online_case_pos =31
complaint_valid_case_pos = 40
complaint_invalid_case_pos = 41
complaint_progress_case_pos = 42
ps_ctrl_hotline_case = 45

cfs_sales_force_dict = {'Year': year,month_pos:'null' ,'channel': channel ,ce_email_case_pos:'null', ce_counter_case_pos:'null',
ce_hotline_case_pos:'null',ce_online_case_pos:'null',
complaint_valid_case_pos:'null',
complaint_invalid_case_pos:'null',complaint_progress_case_pos:'null',
ps_ctrl_hotline_case:'null'
} 

#top3
ORDER_NUMBER_pos = 11
TOP_TR_pos = 12
TR_ENQUIRY_COUNT_pos = 15
TOP_CUST_TYPE_pos = 16
CUST_TYPE_ENQUIRY_COUNT = 19

cfs_top3_dict = {'Year': year,'month':tar_month ,'Year_month': year + YearMonth(tar_month) ,ORDER_NUMBER_pos:'null',
TOP_TR_pos:'null',TR_ENQUIRY_COUNT_pos:'null',
TOP_CUST_TYPE_pos:'null',CUST_TYPE_ENQUIRY_COUNT:'null',
'CHANNEL':channel
} 

#sod
SERVICING_STAFF_pos = 14
SE_EMAIL_CASE_pos = 19
SE_SERVICE_LOG_CASE_pos = 20
SE_HOTLINE_CASE_pos = 21
PS_CTRL_WRITTEN_CASE_pos = 46
PS_CTRL_SURVEY_pos = 50
PS_CTRL_RATING_pos =51

cfs_sod_dict = {'Year': year,month_pos:'null' , 'CHANNEL':channel ,
'LICENSED_CONSULTANT_COUNT':'null','OPERATION_STAFF':'null',SERVICING_STAFF_pos:'null',
SE_EMAIL_CASE_pos:'null',SE_SERVICE_LOG_CASE_pos:'null',SE_HOTLINE_CASE_pos:'null',
PS_CTRL_WRITTEN_CASE_pos:'null',PS_CTRL_SURVEY_pos:'null',PS_CTRL_RATING_pos:'null'
} 

#Operation Staff
cfs_OPERATION_STAFF_pos = 14
cfs_OPERATION_STAFF_dict = {1: 'null',2:'null',
3:'null',4:'null',5:'null',
6:'null',7:'null',8:'null',9:'null',10:'null',11:'null',12:'null'
} 

#Perform
channel = "'Perform'"

ce_email_case_pos = 42
ce_counter_case_pos =44
ce_hotline_case_pos =46
ce_online_case_pos =0
complaint_valid_case_pos = 77
complaint_invalid_case_pos = 78
complaint_progress_case_pos = 79
ps_ctrl_hotline_case = 0
##

op_sales_force_dict = {'Year': year,month_pos:'null' ,'channel': channel , 
ce_email_case_pos:'null', 
ce_counter_case_pos:'null',
ce_hotline_case_pos:'null',
'ce_online_case_pos':'null',
complaint_valid_case_pos:'null',
complaint_invalid_case_pos:'null',
complaint_progress_case_pos:'null',
'ps_ctrl_hotline_case':'null'
} 

#top3
ORDER_NUMBER_pos = 11
TOP_TR_pos = 12
TR_ENQUIRY_COUNT_pos = 15
TOP_CUST_TYPE_pos = 16
CUST_TYPE_ENQUIRY_COUNT = 19

op_top3_dict = {'Year': year,'month':tar_month ,'Year_month': year + YearMonth(tar_month) ,ORDER_NUMBER_pos:'null',
TOP_TR_pos:'null',TR_ENQUIRY_COUNT_pos:'null',
TOP_CUST_TYPE_pos:'null',CUST_TYPE_ENQUIRY_COUNT:'null',
'CHANNEL':"'OWM'"
} 

#sod
perform_SERVICING_STAFF_pos = 14
perform_SE_EMAIL_CASE_pos = 0
perform_SE_Saturn_SERVICE_LOG_CASE_pos = 22
perform_SE_Arch_SERVICE_LOG_CASE_pos = 23
perform_SE_HOTLINE_CASE_pos = 20
perform_PS_CTRL_WRITTEN_CASE_pos = 0
perform_PS_CTRL_SURVEY_pos = 0
perform_PS_CTRL_RATING_pos =0
op_sod_dict = {'Year': year,month_pos:'null' , 'CHANNEL':channel ,
'LICENSED_CONSULTANT_COUNT':'null',
'OPERATION_STAFF':'null',
perform_SERVICING_STAFF_pos:'null',
perform_SE_EMAIL_CASE_pos:'null',
perform_SE_Saturn_SERVICE_LOG_CASE_pos:'null',
perform_SE_Arch_SERVICE_LOG_CASE_pos:'null','SE_HOTLINE_CASE_pos':'null',
'PS_CTRL_WRITTEN_CASE_pos':'null','PS_CTRL_SURVEY_pos':'null','PS_CTRL_RATING_pos':'null'
} 

#Operation Staff
owm_OPERATION_STAFF_pos = 14
owm_OPERATION_STAFF_dict = {1: 'null',2:'null',
3:'null',4:'null',5:'null',
6:'null',7:'null',8:'null',9:'null',10:'null',11:'null',12:'null'
} 

#LICENSED CONSULTANT
CFS_licensed_consultant_dict = {1:'null',2:'null',
3:'null',4:'null',5:'null',
6:'null',7:'null',8:'null',9:'null',10:'null',11:'null',12:'null'
} 

Perform_licensed_consultant_dict = {1:'null',2:'null',
3:'null',4:'null',5:'null',
6:'null',7:'null',8:'null',9:'null',10:'null',11:'null',12:'null'
}

Focus_licensed_consultant_dict = {1:'null',2:'null',
3:'null',4:'null',5:'null',
6:'null',7:'null',8:'null',9:'null',10:'null',11:'null',12:'null'
} 

IFAA_licensed_consultant_dict = { 1:'null',2:'null',
3:'null',4:'null',5:'null',
6:'null',7:'null',8:'null',9:'null',10:'null',11:'null',12:'null'
} 

IFAB_licensed_consultant_dict = { 1:'null',2:'null',
3:'null',4:'null',5:'null',
6:'null',7:'null',8:'null',9:'null',10:'null',11:'null',12:'null'
} 

IFAC_licensed_consultant_dict = { 1:'null',2:'null',
3:'null',4:'null',5:'null',
6:'null',7:'null',8:'null',9:'null',10:'null',11:'null',12:'null'
} 

FTB_licensed_consultant_dict = { 1:'null',2:'null',
3:'null',4:'null',5:'null',
6:'null',7:'null',8:'null',9:'null',10:'null',11:'null',12:'null'
} 

#Special case for IFAA ,focus  

#saleforce
channel = "'IFAA'"
month_pos = 12
ce_email_case_pos = 39
ce_counter_case_pos = 0
ce_hotline_case_pos = 0
ce_online_case_pos =0
complaint_valid_case_pos = 81
complaint_invalid_case_pos = 82
complaint_progress_case_pos = 83
ps_ctrl_hotline_case = 0

ifaa_sales_force_dict = {'Year': year,month_pos:'null' ,'channel': channel ,
ce_email_case_pos:'null', 
'ce_counter_case_pos':'null',
'ce_hotline_case_pos':'null',
'ce_online_case_pos':'null',
complaint_valid_case_pos:'null',
complaint_invalid_case_pos:'null',
complaint_progress_case_pos:'null',
'ps_ctrl_hotline_case':'null'
} 

IFAA_SE_SERVICE_LOG_CASE_pos = 24
SERVICING_STAFF_pos = 14
ifaa_sod_dict ={'Year': year,month_pos:'null' , 'CHANNEL':"'IFAA'" ,
'LICENSED_CONSULTANT_COUNT':'','OPERATION_STAFF':'',SERVICING_STAFF_pos:'null',
'SE_EMAIL_CASE_pos':'null',IFAA_SE_SERVICE_LOG_CASE_pos:'null','SE_HOTLINE_CASE_pos':'null',
'PS_CTRL_WRITTEN_CASE_pos':'null','PS_CTRL_SURVEY_pos':'null','PS_CTRL_RATING_pos':'null'
} 

#saleforce
channel = "'IFAB'"
month_pos = 12
ce_email_case_pos = 40
ce_counter_case_pos =0
ce_hotline_case_pos =0
ce_online_case_pos =0
complaint_valid_case_pos = 85
complaint_invalid_case_pos = 86
complaint_progress_case_pos = 87
ps_ctrl_hotline_case = 0

ifab_sales_force_dict = {'Year': year,month_pos:'null' ,'channel': channel ,
ce_email_case_pos:'null', 
'ce_counter_case_pos':'null',
'ce_hotline_case_pos':'null',
'ce_online_case_pos':'null',
complaint_valid_case_pos:'null',
complaint_invalid_case_pos:'null',
complaint_progress_case_pos:'null',
'ps_ctrl_hotline_case':'null'
} 

IFAB_SE_SERVICE_LOG_CASE_pos = 25
SERVICING_STAFF_pos = 14
ifab_sod_dict ={'Year': year,month_pos:'null' , 'CHANNEL':"'IFAB'" ,
'LICENSED_CONSULTANT_COUNT':'','OPERATION_STAFF':'',SERVICING_STAFF_pos:'null',
'SE_EMAIL_CASE_pos':'null',IFAB_SE_SERVICE_LOG_CASE_pos:'null','SE_HOTLINE_CASE_pos':'null',
'PS_CTRL_WRITTEN_CASE_pos':'null','PS_CTRL_SURVEY_pos':'null','PS_CTRL_RATING_pos':'null'
} 

#saleforce
channel = "'IFAC'"
month_pos = 12
ce_email_case_pos = 41
ce_counter_case_pos =0
ce_hotline_case_pos =0
ce_online_case_pos =0
complaint_valid_case_pos = 89
complaint_invalid_case_pos = 90
complaint_progress_case_pos = 91
ps_ctrl_hotline_case = 0

ifac_sales_force_dict = {'Year': year,month_pos:'null' ,'channel': channel ,
ce_email_case_pos:'null', 
'ce_counter_case_pos':'null',
'ce_hotline_case_pos':'null',
'ce_online_case_pos':'null',
complaint_valid_case_pos:'null',
complaint_invalid_case_pos:'null',
complaint_progress_case_pos:'null',
'ps_ctrl_hotline_case':'null'
} 

IFAC_SE_SERVICE_LOG_CASE_pos = 26
SERVICING_STAFF_pos = 14
ifac_sod_dict ={'Year': year,month_pos:'null' , 'CHANNEL':"'IFAC'" ,
'LICENSED_CONSULTANT_COUNT':'','OPERATION_STAFF':'',SERVICING_STAFF_pos:'null',
'SE_EMAIL_CASE_pos':'null',IFAC_SE_SERVICE_LOG_CASE_pos:'null','SE_HOTLINE_CASE_pos':'null',
'PS_CTRL_WRITTEN_CASE_pos':'null','PS_CTRL_SURVEY_pos':'null','PS_CTRL_RATING_pos':'null'
} 

#saleforce
channel = "'FTB'"
month_pos = 12
ce_email_case_pos = 41
ce_counter_case_pos =0
ce_hotline_case_pos =0
ce_online_case_pos =0
complaint_valid_case_pos = 89
complaint_invalid_case_pos = 90
complaint_progress_case_pos = 91
ps_ctrl_hotline_case = 0

ftb_sales_force_dict = {'Year': year,month_pos:'null' ,'channel': channel ,
ce_email_case_pos:'null', 
'ce_counter_case_pos':'null',
'ce_hotline_case_pos':'null',
'ce_online_case_pos':'null',
complaint_valid_case_pos:'null',
complaint_invalid_case_pos:'null',
complaint_progress_case_pos:'null',
'ps_ctrl_hotline_case':'null'
} 

FTB_SE_SERVICE_LOG_CASE_pos = 26
SERVICING_STAFF_pos = 14
ftb_sod_dict ={'Year': year,month_pos:'null' , 'CHANNEL':"'FTB'" ,
'LICENSED_CONSULTANT_COUNT':'','OPERATION_STAFF':'',SERVICING_STAFF_pos:'null',
'SE_EMAIL_CASE_pos':'null',FTB_SE_SERVICE_LOG_CASE_pos:'null','SE_HOTLINE_CASE_pos':'null',
'PS_CTRL_WRITTEN_CASE_pos':'null','PS_CTRL_SURVEY_pos':'null','PS_CTRL_RATING_pos':'null'
} 

#saleforce
channel = "'FOCUS'"
month_pos = 12
ce_email_case_pos = 38
ce_counter_case_pos =0
ce_hotline_case_pos =0
ce_online_case_pos =0
complaint_valid_case_pos = 73
complaint_invalid_case_pos = 74
complaint_progress_case_pos = 75
ps_ctrl_hotline_case = 0

SERVICING_STAFF_pos = 14
focus_sales_force_dict = {'Year': year,month_pos:'null' ,'channel': channel ,
ce_email_case_pos:'null', 
'ce_counter_case_pos':'null',
'ce_hotline_case_pos':'null',
'ce_online_case_pos':'null',
complaint_valid_case_pos:'null',
complaint_invalid_case_pos:'null',
complaint_progress_case_pos:'null',
'ps_ctrl_hotline_case':'null'
} 


focus_SE_SERVICE_LOG_CASE_pos = 21
SERVICING_STAFF_pos = 14
#Special case for IFAA ,focus  
focus_sod_dict ={'Year': year,month_pos:'' , 'CHANNEL':"'FOCUS'" ,
'LICENSED_CONSULTANT_COUNT':'','OPERATION_STAFF':'',SERVICING_STAFF_pos:'null',
'SE_EMAIL_CASE_pos':'null',focus_SE_SERVICE_LOG_CASE_pos:'null','SE_HOTLINE_CASE_pos':'null',
'PS_CTRL_WRITTEN_CASE_pos':'null','PS_CTRL_SURVEY_pos':'null','PS_CTRL_RATING_pos':'null'
} 

def checkNone(value) :
    if value == None or str(value).strip() == '-': 
        return "Null"
    else:
        return value
    
    
def checkNoneReturn0(value) :
    if value == 'Null' or value == '-': 
        return 0
    else:
        return value
    
def insertDB(table_name,dict):
     values_string = '('+','.join(map(str,dict.values()))+')'    
     sql = """INSERT INTO %s
          VALUES %s"""%(table_name,values_string)
     print(sql)
     with pyodbc.connect('DRIVER='+driver+';SERVER='+server+';PORT=1433;DATABASE='+database+';UID='+dbusername+';PWD='+ dbpassword) as conn:
          with conn.cursor() as cursor:
               cursor.execute(sql)
               
def deleteDBByYear(table_name ,year):
    with pyodbc.connect('DRIVER='+driver+';SERVER='+server+';PORT=1433;DATABASE='+database+';UID='+dbusername+';PWD='+ dbpassword) as conn:
        with conn.cursor() as cursor:
            sql = "delete from " + table_name + " where year =" + year
            print(sql)
            cursor.execute(sql)

def deleteDBByYearMonth(table_name ,year,month):
    with pyodbc.connect('DRIVER='+driver+';SERVER='+server+';PORT=1433;DATABASE='+database+';UID='+dbusername+';PWD='+ dbpassword) as conn:
        with conn.cursor() as cursor:
            sql = "delete from " + table_name + " where year =" + year + " and month = " + month
            print(sql)
            cursor.execute(sql)
            


#Saleforce
# deleteDBByYear(sales_force_table_name ,year)
#CFS
for col_cells in CFS_sheet.iter_cols(saleforce_start_column, saleforce_end_column):
    for cell in col_cells:
        for k, v in cfs_sales_force_dict.items():
            if(cell.row == k):
                if (cell.row ==month_pos):
                    month_name = cell.value
                    datetime_object = datetime.datetime.strptime(month_name, "%b")
                    cfs_sales_force_dict[k] = datetime_object.month
                else:    
                    cfs_sales_force_dict[k] =  checkNone(cell.value)
    # print("hello sk2 " + sales_force_table_name,cfs_sales_force_dict)
    insertDB(sales_force_table_name,cfs_sales_force_dict)


#Perform
for col_cells in OP_sheet.iter_cols(saleforce_start_column, saleforce_end_column):
    for cell in col_cells:
        for k, v in op_sales_force_dict.items():
            if(cell.row == k):
                if (cell.row ==month_pos):
                    month_name = cell.value
                    datetime_object = datetime.datetime.strptime(month_name, "%b")
                    op_sales_force_dict[k] = datetime_object.month
                else:    
                    op_sales_force_dict[k] =  checkNone(cell.value)

    insertDB(sales_force_table_name,op_sales_force_dict)


#IFAA
for col_cells in OP_sheet.iter_cols(saleforce_start_column, saleforce_end_column):
    for cell in col_cells:
        for k, v in ifaa_sales_force_dict.items():
            if(cell.row == k):
                if (cell.row ==month_pos):
                    month_name = cell.value
                    datetime_object = datetime.datetime.strptime(month_name, "%b")
                    ifaa_sales_force_dict[k] = datetime_object.month
                else:    
                    ifaa_sales_force_dict[k] =  checkNone(cell.value)

    insertDB(sales_force_table_name,ifaa_sales_force_dict)


#IFAB
for col_cells in OP_sheet.iter_cols(saleforce_start_column, saleforce_end_column):
    for cell in col_cells:
        for k, v in ifab_sales_force_dict.items():
            if(cell.row == k):
                if (cell.row ==month_pos):
                    month_name = cell.value
                    datetime_object = datetime.datetime.strptime(month_name, "%b")
                    ifab_sales_force_dict[k] = datetime_object.month
                else:    
                    ifab_sales_force_dict[k] =  checkNone(cell.value)

    insertDB(sales_force_table_name,ifab_sales_force_dict)

#IFAC
for col_cells in OP_sheet.iter_cols(saleforce_start_column, saleforce_end_column):
    for cell in col_cells:
        for k, v in ifac_sales_force_dict.items():
            if(cell.row == k):
                if (cell.row ==month_pos):
                    month_name = cell.value
                    datetime_object = datetime.datetime.strptime(month_name, "%b")
                    ifac_sales_force_dict[k] = datetime_object.month
                else:    
                    ifac_sales_force_dict[k] =  checkNone(cell.value)

    insertDB(sales_force_table_name,ifac_sales_force_dict)

#FTB
for col_cells in OP_sheet.iter_cols(saleforce_start_column, saleforce_end_column):
    for cell in col_cells:
        for k, v in ftb_sales_force_dict.items():
            if(cell.row == k):
                if (cell.row ==month_pos):
                    month_name = cell.value
                    datetime_object = datetime.datetime.strptime(month_name, "%b")
                    ftb_sales_force_dict[k] = datetime_object.month
                else:    
                    ftb_sales_force_dict[k] =  checkNone(cell.value)

    insertDB(sales_force_table_name,ftb_sales_force_dict)
    
#Focus
for col_cells in OP_sheet.iter_cols(saleforce_start_column, saleforce_end_column):
    for cell in col_cells:
        for k, v in focus_sales_force_dict.items():
            if(cell.row == k):
                if (cell.row ==month_pos):
                    month_name = cell.value
                    datetime_object = datetime.datetime.strptime(month_name, "%b")
                    focus_sales_force_dict[k] = datetime_object.month
                else:    
                    focus_sales_force_dict[k] =  checkNone(cell.value)

    insertDB(sales_force_table_name,focus_sales_force_dict)
    
    
#Top3
deleteDBByYearMonth(top3_table_name ,year,str(tar_month))
#CFS
for col_cells in CFS_sheet.iter_rows(top3_start_row, top3_end_row):
    for cell in col_cells:
        if (cell.value != None):
            for k, v in cfs_top3_dict.items():
                if(cell.column == k):
                    cfs_top3_dict[k] =  "'" + str(cell.value) + "'"
    insertDB(top3_table_name,cfs_top3_dict) 

#perform
for col_cells in OP_sheet.iter_rows(top3_start_row, top3_end_row):
    for cell in col_cells:
        if (cell.value != None):
            for k, v in op_top3_dict.items():
                if(cell.column == k):
                    op_top3_dict[k] =  "'" + str(cell.value) + "'"
    insertDB(top3_table_name,op_top3_dict)     
    

#Get Operation Staff 
OWM_operation_count = 1
for col_cells in OWM_operation_sheet.iter_cols(operation_start_column, operation_end_column):
    for cell in col_cells:
            if(cell.row == owm_OPERATION_STAFF_pos):
                owm_OPERATION_STAFF_dict[OWM_operation_count] =checkNone(cell.value)
                OWM_operation_count = OWM_operation_count +1

print(owm_OPERATION_STAFF_dict)          
            
#Get Licensed Consultant Count 
with pyodbc.connect('DRIVER='+driver+';SERVER=tcp:'+server+';PORT=1433;DATABASE='+database+';UID='+dbusername+';PWD='+ dbpassword) as conn:
    with conn.cursor() as cursor:
        sql = """select Year , cast(month as int ) month ,CHANNEL ,count(distinct employee_key) count_nb from Dwh_Mi_CfsLicensedConsultant where YEAR = %s group by Year , month ,CHANNEL order by CHANNEL ,month"""%(year)
        cursor.execute(sql)
        row = cursor.fetchone()
        while row:
            if (row[2] == 'CFS'):
                 CFS_licensed_consultant_dict[row[1]] = checkNone(row[3])
            if (row[2] == 'Perform'):
                 Perform_licensed_consultant_dict[row[1]] = checkNone(row[3])    
            if (row[2] == 'FOCUS'):
                 Focus_licensed_consultant_dict[row[1]] = checkNone(row[3])
            if (row[2] == 'IFAA'):
                 IFAA_licensed_consultant_dict[row[1]] = checkNone(row[3])
            if (row[2] == 'IFAB'):
                 IFAB_licensed_consultant_dict[row[1]] = checkNone(row[3])
            if (row[2] == 'IFAC'):
                 IFAC_licensed_consultant_dict[row[1]] = checkNone(row[3])
            if (row[2] == 'IFAFTBC'):
                 FTB_licensed_consultant_dict[row[1]] = checkNone(row[3])
            row = cursor.fetchone()

print(CFS_licensed_consultant_dict)
print(Perform_licensed_consultant_dict)
print(Focus_licensed_consultant_dict)
print(IFAA_licensed_consultant_dict)
print(IFAB_licensed_consultant_dict)
print(IFAC_licensed_consultant_dict)

#SOD
deleteDBByYear(sod_table_name ,year)
#CFS
for col_cells in CFS_sheet.iter_cols(saleforce_start_column, saleforce_end_column):
    for cell in col_cells:
        for k, v in cfs_sod_dict.items():
            if(cell.row == k):  
                if (cell.row ==month_pos):
                    month_name = cell.value
                    datetime_object = datetime.datetime.strptime(month_name, "%b")
                    cfs_sod_dict[k] = datetime_object.month
                else:    
                    cfs_sod_dict[k] =  checkNone(cell.value)

    if (cfs_sod_dict['CHANNEL'] == "'CFS'"):
        cfs_sod_dict['LICENSED_CONSULTANT_COUNT'] = checkNone(CFS_licensed_consultant_dict[cfs_sod_dict[month_pos]])
        cfs_sod_dict['OPERATION_STAFF'] = checkNone(owm_OPERATION_STAFF_dict[cfs_sod_dict[month_pos]])
        print("Hello SK " + str(checkNone(CFS_licensed_consultant_dict[cfs_sod_dict[month_pos]])))
        
    insertDB(sod_table_name,cfs_sod_dict)


#Perform
for col_cells in OP_sheet.iter_cols(saleforce_start_column, saleforce_end_column):
    for cell in col_cells:
        for k, v in op_sod_dict.items():
            if(cell.row == k):  
                if (cell.row ==month_pos):
                    month_name = cell.value
                    datetime_object = datetime.datetime.strptime(month_name, "%b")
                    op_sod_dict[k] = datetime_object.month
                else:    
                    op_sod_dict[k] =  checkNone(cell.value)

    if (op_sod_dict['CHANNEL'] == "'Perform'"):
        op_sod_dict['LICENSED_CONSULTANT_COUNT'] = checkNone(Perform_licensed_consultant_dict[op_sod_dict[month_pos]])
        op_sod_dict['OPERATION_STAFF'] = checkNone(owm_OPERATION_STAFF_dict[op_sod_dict[month_pos]])


     #Sepical handle for   Perform Service Log (Saturn) and  Perform Service Log (ARCH)
    op_sod_dict[perform_SE_Saturn_SERVICE_LOG_CASE_pos] = int(checkNoneReturn0(op_sod_dict[perform_SE_Saturn_SERVICE_LOG_CASE_pos])) + int(checkNoneReturn0(op_sod_dict[perform_SE_Arch_SERVICE_LOG_CASE_pos]))
    op_sod_dict_temp = op_sod_dict.copy()
    del op_sod_dict_temp[perform_SE_Arch_SERVICE_LOG_CASE_pos]
    insertDB(sod_table_name,op_sod_dict_temp)

#Focus
for col_cells in OP_sheet.iter_cols(saleforce_start_column, saleforce_end_column):
    for cell in col_cells:
        for k, v in focus_sod_dict.items():
            if(cell.row == k):  
                if (cell.row ==month_pos):
                    month_name = cell.value
                    datetime_object = datetime.datetime.strptime(month_name, "%b")
                    focus_sod_dict[k] = datetime_object.month
                else:    
                    focus_sod_dict[k] =  checkNone(cell.value)

    if (focus_sod_dict['CHANNEL'] == "'FOCUS'"):
        focus_sod_dict['LICENSED_CONSULTANT_COUNT'] = checkNone(Focus_licensed_consultant_dict[focus_sod_dict[month_pos]])
        focus_sod_dict['OPERATION_STAFF'] = checkNone(owm_OPERATION_STAFF_dict[focus_sod_dict[month_pos]])

    insertDB(sod_table_name,focus_sod_dict)

#IFAA
for col_cells in OP_sheet.iter_cols(saleforce_start_column, saleforce_end_column):
    for cell in col_cells:
        for k, v in ifaa_sod_dict.items():
            if(cell.row == k):  
                if (cell.row ==month_pos):
                    month_name = cell.value
                    datetime_object = datetime.datetime.strptime(month_name, "%b")
                    ifaa_sod_dict[k] = datetime_object.month
                else:    
                    ifaa_sod_dict[k] =  checkNone(cell.value)

    if (ifaa_sod_dict['CHANNEL'] == "'IFAA'"):
        ifaa_sod_dict['LICENSED_CONSULTANT_COUNT'] = checkNone(IFAA_licensed_consultant_dict[ifaa_sod_dict[month_pos]])
        ifaa_sod_dict['OPERATION_STAFF'] = checkNone(owm_OPERATION_STAFF_dict[ifaa_sod_dict[month_pos]])

    insertDB(sod_table_name,ifaa_sod_dict)

#IFAB
for col_cells in OP_sheet.iter_cols(saleforce_start_column, saleforce_end_column):
    for cell in col_cells:
        for k, v in ifab_sod_dict.items():
            if(cell.row == k):  
                if (cell.row ==month_pos):
                    month_name = cell.value
                    datetime_object = datetime.datetime.strptime(month_name, "%b")
                    ifab_sod_dict[k] = datetime_object.month
                else:    
                    ifab_sod_dict[k] =  checkNone(cell.value)

    if (ifab_sod_dict['CHANNEL'] == "'IFAB'"):
        ifab_sod_dict['LICENSED_CONSULTANT_COUNT'] = checkNone(IFAB_licensed_consultant_dict[ifab_sod_dict[month_pos]])
        ifab_sod_dict['OPERATION_STAFF'] = checkNone(owm_OPERATION_STAFF_dict[ifab_sod_dict[month_pos]])

    insertDB(sod_table_name,ifab_sod_dict)
    
#IFAC
for col_cells in OP_sheet.iter_cols(saleforce_start_column, saleforce_end_column):
    for cell in col_cells:
        for k, v in ifac_sod_dict.items():
            if(cell.row == k):  
                if (cell.row ==month_pos):
                    month_name = cell.value
                    datetime_object = datetime.datetime.strptime(month_name, "%b")
                    ifac_sod_dict[k] = datetime_object.month
                else:    
                    ifac_sod_dict[k] =  checkNone(cell.value)

    if (ifac_sod_dict['CHANNEL'] == "'IFAC'"):
        ifac_sod_dict['LICENSED_CONSULTANT_COUNT'] = checkNone(IFAC_licensed_consultant_dict[ifac_sod_dict[month_pos]])
        ifac_sod_dict['OPERATION_STAFF'] = checkNone(owm_OPERATION_STAFF_dict[ifac_sod_dict[month_pos]])

    insertDB(sod_table_name,ifac_sod_dict)

#FTB
for col_cells in OP_sheet.iter_cols(saleforce_start_column, saleforce_end_column):
    for cell in col_cells:
        for k, v in ftb_sod_dict.items():
            if(cell.row == k):  
                if (cell.row ==month_pos):
                    month_name = cell.value
                    datetime_object = datetime.datetime.strptime(month_name, "%b")
                    ftb_sod_dict[k] = datetime_object.month
                else:    
                    ftb_sod_dict[k] =  checkNone(cell.value)

    if (ftb_sod_dict['CHANNEL'] == "'FTB'"):
        ftb_sod_dict['LICENSED_CONSULTANT_COUNT'] = checkNone(FTB_licensed_consultant_dict[ftb_sod_dict[month_pos]])
        ftb_sod_dict['OPERATION_STAFF'] = checkNone(owm_OPERATION_STAFF_dict[ftb_sod_dict[month_pos]])

    insertDB(sod_table_name,ftb_sod_dict)

EndDay = calendar.monthrange(last_month.year,last_month.month)[1] 
                
#Create Service Overview
OWM_licensed_consultant_dict = { 1:'null',2:'null',3:'null',4:'null',5:'null',6:'null',7:'null',8:'null',9:'null',10:'null',11:'null',12:'null'} 
with pyodbc.connect('DRIVER='+driver+';SERVER=tcp:'+server+';PORT=1433;DATABASE='+database+';UID='+dbusername+';PWD='+ dbpassword) as conn:
    with conn.cursor() as cursor:
        sql = """select Year , cast(month as int ) month ,LicenseCorp ,count(distinct employee_key) count_nb from Dwh_Mi_CfsLicensedConsultant a left join convoyProduct.dbo.PrdDistributor b on a.CHANNEL=b.DistributorCode where YEAR = %s group by Year , month ,LicenseCorp"""%(year)
        cursor.execute(sql)
        for row in cursor.fetchall(): 
            if (row[2] == 'OWM'):
                OWM_licensed_consultant_dict[row[1]] = checkNone(row[3])

print(OWM_licensed_consultant_dict)
print(CFS_licensed_consultant_dict)

#Get report sent date
with pyodbc.connect('DRIVER='+driver+';SERVER=tcp:'+server+';PORT=1433;DATABASE='+database+';UID='+dbusername+';PWD='+ dbpassword) as conn:
    with conn.cursor() as cursor:
        sql = """select date_id from Dwh_D_Date where MI_FIRST_SENT_DATE is not null and YEARMONTH = %s"""%(datetime.datetime.today().strftime('%Y%m'))
        cursor.execute(sql)
        mi_first_sent_date = cursor.fetchval().strftime('%Y%m%d')
        print(mi_first_sent_date)

folder_path_output=serv_blob_path #+ 'output/'+yyyymm+'/'
serv_report_name = mi_first_sent_date + '_'+ 'MI Serv Mgmt Overview.xlsx' #'MI Serv Mgmt Overview - '+ yyyy_mm +'.xlsx'

### save to service_management_complete
mgmt_wb_obj = download_excel_file(mgmt_folder_url, "MI Serv Mgmt Overview")

if mgmt_wb_obj and 'CFS' in mgmt_wb_obj.sheetnames:
    ws_CFS = mgmt_wb_obj['CFS']
    ws_OP = mgmt_wb_obj["OP"]

# Sheet CFS
r_cfs = 13  # start from row13
c_cfs = 5  # col E
# ws_CFS = wb['CFS'] 
for key, rows in CFS_licensed_consultant_dict.items():
    if rows != 'null':
        ws_CFS.cell(row=r_cfs, column=c_cfs).value = rows
    c_cfs += 1

# Sheet OP
r_op = 13  # start from row13
c_op = 5  # col E
# ws_OP = wb['OP'] 
for key, rows in OWM_licensed_consultant_dict.items():
    if rows != 'null':
        ws_OP.cell(row=r_op, column=c_op).value = rows
    c_op += 1

# Save the modified workbook to a BytesIO object
modified_excel = io.BytesIO()
mgmt_wb_obj.save(modified_excel)
modified_excel.seek(0)  # Reset the pointer to the beginning of the stream

serv_report_name = mi_first_sent_date + '_'+ 'MI Serv Mgmt Overview.xlsx' #'MI Serv Mgmt Overview - '+ yyyy_mm +'.xlsx'

upload_folder_url = "Shared%20Documents/Shared/01_MI Reports/Raw/Service Management Complete/"

serv_report_name = f"{mi_first_sent_date}_MI Serv Mgmt Overview.xlsx"
target_file_url = f"{upload_folder_url}{serv_report_name}"
target_folder = ctx.web.get_folder_by_server_relative_url(upload_folder_url)
uploaded_file = target_folder.upload_file(serv_report_name, modified_excel).execute_query()

# Print confirmation
print(f"File uploaded successfully to {upload_folder_url}{serv_report_name}")


            
      
           
